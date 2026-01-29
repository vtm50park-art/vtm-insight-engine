import streamlit as st
import pandas as pd
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import sys
from datetime import datetime, timedelta
import io
from openpyxl.styles import Font, PatternFill
import time
import json
import os

# plotly 안전 임포트
try:
    import plotly.express as px
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

# ==================== 설정 ====================
st.set_page_config(page_title="VTM Insight V3.0", layout="wide")

API_KEYS = [
    'AIzaSyCcd7eue6-OcXSw5v_40kEwV_fJ65M8mqE',
    'AIzaSyC9M9RuyJih1MaEdKyUy7m2cx-_7r8zBr8',
    'AIzaSyD-4PTGLzaa07HFaUHCjcM3cvXPSS_W7Nk'
]

if 'active_idx' not in st.session_state:
    st.session_state['active_idx'] = 0
if 'vtm_user' not in st.session_state:
    st.session_state['vtm_user'] = None
if 'vtm_df' not in st.session_state:
    st.session_state['vtm_df'] = None
if 'failed_keys' not in st.session_state:
    st.session_state['failed_keys'] = set()
if 'access_logs' not in st.session_state:
    st.session_state['access_logs'] = []

# ==================== 로그 관리 ====================
LOG_FILE = "vtm_access_logs.json"

def save_log(user, action):
    """로그 저장"""
    log_entry = {
        "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "user": user,
        "action": action
    }
    
    # 세션 로그에 추가
    st.session_state['access_logs'].append(log_entry)
    
    # 파일에 저장 (영구 보관)
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                logs = json.load(f)
        else:
            logs = []
        
        logs.append(log_entry)
        
        with open(LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
    except:
        pass  # 파일 저장 실패해도 계속 진행

def load_logs():
    """로그 불러오기"""
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []
    except:
        return []

def create_log_excel(logs):
    """로그를 엑셀로 변환"""
    df = pd.DataFrame(logs)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="접속기록")
        worksheet = writer.sheets["접속기록"]
        
        # 열 너비 조정
        worksheet.column_dimensions['A'].width = 20  # timestamp
        worksheet.column_dimensions['B'].width = 15  # user
        worksheet.column_dimensions['C'].width = 50  # action
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    output.seek(0)
    return output.getvalue()

# ==================== 채널 정보 ====================
CHANNELS = {
    "브이티엠 엔터": {
        "desc": "힙합 및 AI 뮤직비디오 제작 트렌드 분석",
        "purpose": "힙합 플레이리스트의 제목, 썸네일, 곡 구성을 분석하여 브이티엠만의 힙합 플리 제작. 경쟁 플리의 조회수 패턴과 알고리즘 최적화 전략 벤치마킹.",
        "insight": "떡상 플리는 '장르 믹스 + 무드 키워드(심야, 드라이브, 작업용)' 조합이 핵심. 썸네일은 고대비 색상 + 큰 장르명 필수.",
        "hiphop": {"q": "힙합 플레이리스트 rap playlist", "region": "KR"},
        "ai": {"q": "AI Generated Music Video Sora Veo Runway", "region": ""}
    },
    "시니어 스마일": {
        "desc": "실버 세대 맞춤형 트로트 큐레이션",
        "purpose": "시니어 세대를 위한 트로트 플리 썸네일(고대비, 큰 글씨, 따뜻한 색감) 및 제목 전략 벤치마킹.",
        "insight": "시니어는 '추억, 향수, 정겨운' 키워드에 반응. 썸네일은 노란색/빨간색 계열 + 명조체 큰 글씨.",
        "boksun": {"q": "신나는 트로트 메들리", "region": "KR"},
        "youngwon": {"q": "애절한 트로트", "region": "KR"}
    },
    "키즈타임": {
        "desc": "Cocomelon 기반 영유아 교육 콘텐츠",
        "purpose": "영유아 교육 콘텐츠의 색감, 캐릭터 디자인, 반복 패턴 분석.",
        "insight": "키즈 콘텐츠는 '반복+리듬감+밝은 색상'이 생명. 제목에 숫자 명시 필수.",
        "q": "Kids Songs Nursery Rhymes",
        "region": "US"
    },
    "VTM ADLAB": {
        "desc": "글로벌 브랜드 광고 분석",
        "purpose": "최신 바이럴 광고 캠페인 분석.",
        "insight": "바이럴 광고는 '감동+반전+공감' 3박자. 첫 3초가 생명.",
        "q": "Brand Commercial Viral",
        "region": "US"
    },
    "SOUND BALM": {
        "desc": "숙면 유도 음악",
        "purpose": "해외 숙면 음악 채널 전략 분석.",
        "insight": "숙면 콘텐츠는 '시간 길이'가 핵심. 8시간 이상 영상이 알고리즘 우대.",
        "q": "Sleep Music 8 Hours",
        "region": "US"
    },
    "EUN2 채널": {
        "desc": "로파이 음악",
        "purpose": "해외 로파이 채널 전략 분석.",
        "insight": "로파이는 '24/7 라이브 스트림' 포맷이 강력.",
        "q": "Lofi hip hop beats",
        "region": "US"
    },
    "시티팝 채널": {
        "desc": "레트로 시티팝",
        "purpose": "80년대 레트로 감성 분석.",
        "insight": "시티팝은 '일본어 제목 + 80s 명시'가 핵심.",
        "q": "City Pop 80s Japanese",
        "region": "US"
    },
    "케이엠코스랩": {
        "desc": "화장품 OEM/ODM",
        "purpose": "화장품 제조 콘텐츠 분석.",
        "insight": "화장품 B2B 콘텐츠는 '투명성+전문성'이 신뢰도 결정.",
        "q": "Cosmetic Manufacturing OEM",
        "region": "KR"
    },
    "다이어트 팩트": {
        "desc": "과학적 다이어트 정보",
        "purpose": "다이어트 정보 전달 전략 분석.",
        "insight": "다이어트 콘텐츠는 '구체적 숫자'가 클릭 유도.",
        "q": "다이어트 식단 전문가",
        "region": "KR"
    },
    "리커버 채널": {
        "desc": "AI 리커버",
        "purpose": "AI 커버 영상 전략 분석.",
        "insight": "AI 커버는 '유명 아티스트 음색 변환'이 핵심.",
        "q": "AI Cover Korean Song",
        "region": "US"
    }
}

# ==================== 엑셀 생성 ====================
def create_vtm_excel(df, sheet_name, user_name):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df = df.drop(columns=['썸네일']) if '썸네일' in df.columns else df
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
        
        for row in range(2, min(5, len(export_df) + 2)):
            for col in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                if row == 2:
                    cell.fill = gold_fill
                elif row == 3:
                    cell.fill = silver_fill
                elif row == 4:
                    cell.fill = bronze_fill
        
        for row_idx in range(2, len(export_df) + 2):
            if "영상 주소" in export_df.columns:
                video_url_col = export_df.columns.get_loc("영상 주소") + 1
                cell = worksheet.cell(row=row_idx, column=video_url_col)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
            
            if "채널 주소" in export_df.columns:
                channel_url_col = export_df.columns.get_loc("채널 주소") + 1
                cell = worksheet.cell(row=row_idx, column=channel_url_col)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
        
        for i in range(len(export_df.columns)):
            worksheet.column_dimensions[chr(65 + i)].width = 25
        
        last_row = len(export_df) + 3
        info_text = f"모든 분석 데이터의 저작권은 (주)브이티엠 에 있습니다. (개발자: 박동진 본부장 / 분석자: {user_name})"
        cell = worksheet.cell(row=last_row, column=1, value=info_text)
        cell.font = Font(bold=True, color="000000")
    
    output.seek(0)
    return output.getvalue()

def log_vtm(user, action):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{now}] [{user}] {action}")
    sys.stdout.flush()
    save_log(user, action)

def get_youtube_client():
    for attempt in range(len(API_KEYS)):
        idx = st.session_state['active_idx']
        if idx in st.session_state['failed_keys']:
            st.session_state['active_idx'] = (idx + 1) % len(API_KEYS)
            continue
        try:
            return build('youtube', 'v3', developerKey=API_KEYS[idx])
        except:
            st.session_state['failed_keys'].add(idx)
            st.session_state['active_idx'] = (idx + 1) % len(API_KEYS)
    return None

# ==================== YouTube 검색 ====================
def fetch_youtube_native(query, region, order_type, period, content_type, max_results=20):
    youtube = get_youtube_client()
    if youtube is None:
        return None
    
    published_after = None
    if period == "이번 주":
        published_after = (datetime.now() - timedelta(days=7)).isoformat() + "Z"
    elif period == "이번 달":
        published_after = (datetime.now() - timedelta(days=30)).isoformat() + "Z"
    
    if content_type == "쇼츠만":
        video_duration = "short"
        search_query = query + " shorts"
    elif content_type == "미드폼만":
        video_duration = "medium"
        search_query = query
    else:
        video_duration = None
        search_query = query
    
    v_list = []
    attempts = 0
    
    while attempts < len(API_KEYS) * 2:
        try:
            search_params = {
                "q": search_query,
                "part": "snippet",
                "type": "video",
                "maxResults": max_results,
                "order": order_type
            }
            
            if region:
                search_params["regionCode"] = region
            
            if published_after:
                search_params["publishedAfter"] = published_after
            if video_duration:
                search_params["videoDuration"] = video_duration
            
            search_res = youtube.search().list(**search_params).execute()
            
            if not search_res.get('items'):
                return None
            
            for item in search_res.get('items', []):
                v_id = item['id'].get('videoId')
                if not v_id:
                    continue
                
                try:
                    v_res = youtube.videos().list(
                        part="statistics,snippet",
                        id=v_id
                    ).execute()
                    
                    if not v_res.get('items'):
                        continue
                    
                    v_info = v_res['items'][0]
                    c_id = v_info['snippet']['channelId']
                    
                    c_res = youtube.channels().list(
                        part="statistics,snippet",
                        id=c_id
                    ).execute()
                    
                    if not c_res.get('items'):
                        continue
                    
                    c_info = c_res['items'][0]
                    subs = int(c_info['statistics'].get('subscriberCount', 0))
                    
                    if subs < 1:
                        continue
                    
                    views = int(v_info['statistics'].get('viewCount', 0))
                    likes = int(v_info['statistics'].get('likeCount', 0))
                    comments = int(v_info['statistics'].get('commentCount', 0))
                    
                    viral_score = round((views / subs) * 100, 2) if subs > 0 else 0
                    
                    if viral_score < 50:
                        continue
                    
                    god_badge = "🔥신의 간택" if viral_score >= 10000 else ""
                    
                    thumbnail_url = v_info['snippet']['thumbnails'].get('medium', {}).get('url', '')
                    channel_created = c_info['snippet'].get('publishedAt', '')[:10]
                    
                    v_list.append({
                        "썸네일": thumbnail_url,
                        "채널명": v_info['snippet']['channelTitle'],
                        "동영상 제목": v_info['snippet']['title'] + (" " + god_badge if god_badge else ""),
                        "업로드 일자": v_info['snippet']['publishedAt'][:10],
                        "채널 개설일": channel_created,
                        "조회수": views,
                        "구독자 수": subs,
                        "Viral Score": viral_score,
                        "구독자 대비 조회": f"{viral_score/100:.2f}x",
                        "좋아요율": f"{(likes/views*100):.1f}%",
                        "좋아요 수": likes,
                        "댓글 수": comments,
                        "영상 주소": f"https://youtu.be/{v_id}",
                        "채널 주소": f"https://www.youtube.com/channel/{c_id}"
                    })
                    
                except HttpError as e:
                    if e.resp.status == 403:
                        raise
                    continue
            
            return v_list if v_list else None
            
        except HttpError as e:
            if e.resp.status == 403:
                current_key = st.session_state['active_idx']
                st.session_state['failed_keys'].add(current_key)
                st.session_state['active_idx'] = (current_key + 1) % len(API_KEYS)
                youtube = get_youtube_client()
                if youtube is None:
                    return None
                time.sleep(1)
                attempts += 1
                continue
    
    return None

# ==================== 개별 영상 분석 ====================
def analyze_video(row, rank, channel_name):
    engagement_rate = (row['좋아요 수'] + row['댓글 수']) / row['조회수'] * 100
    viral_level = "신의 간택" if row['Viral Score'] >= 10000 else "초대박" if row['Viral Score'] >= 1000 else "성공"
    
    analysis = f"""
### 🔥 떡상 분석

**📊 성과 지표**
- 조회수: {row['조회수']:,}회 | 구독자: {row['구독자 수']:,}명
- Viral Score: {row['Viral Score']:,.0f}% ({viral_level})
- 구독자 대비: {row['Viral Score']/100:.1f}배 | 참여율: {engagement_rate:.2f}%

**💡 떡상 핵심 요인**
"""
    
    title = row['동영상 제목'].replace("🔥신의 간택", "").strip()
    
    if any(x in title.lower() for x in ['playlist', '플레이리스트', 'mix', 'compilation']):
        analysis += "✅ **플레이리스트 포맷**: 연속 재생으로 시청 시간 극대화\n"
    if any(char.isdigit() for char in title):
        analysis += "✅ **숫자 포함**: 구체적 정보 제공으로 클릭 유도\n"
    
    if row['Viral Score'] >= 10000:
        analysis += "✅ **알고리즘 최적화**: 홈 피드 대량 노출 + 바이럴 확산\n"
    elif row['Viral Score'] >= 1000:
        analysis += "✅ **추천 알고리즘 탑승**: 관련 영상 추천 적극 노출\n"
    
    if engagement_rate > 3:
        analysis += "✅ **높은 참여율**: 충성 팬층 확보로 알고리즘 신호 강화\n"
    
    if row['구독자 수'] < 10000:
        analysis += "✅ **소규모 채널 성공**: 틈새 시장 정확한 타겟팅\n"
    
    analysis += f"""
**🎯 {channel_name} 적용 인사이트**
{CHANNELS[channel_name]['insight']}
"""
    
    return analysis

# ==================== 로그인 ====================
if st.session_state['vtm_user'] is None:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<h1 style='text-align: center;'>🛡️ (주)브이티엠</h1>", unsafe_allow_html=True)
        st.markdown("<h2 style='text-align: center;'>인사이트 엔진 V3.0</h2>", unsafe_allow_html=True)
        st.markdown("---")
        
        user_choice = st.selectbox("👤 담당자", 
                                   ["선택하세요", "박동진 본부장", "김해정 팀장", "강유미 대리", "김소원 주임", "이하림 총무"])
        
        if user_choice == "박동진 본부장":
            pw = st.text_input("🔐 비밀번호", type="password")
            if st.button("🚀 시스템 가동", use_container_width=True, type="primary"):
                if pw == "5638":
                    st.session_state['vtm_user'] = user_choice
                    save_log(user_choice, "시스템 접속 (관리자)")
                    st.rerun()
                else:
                    st.error("❌ 비밀번호 오류")
        elif user_choice != "선택하세요":
            if st.button("🚀 시스템 가동", use_container_width=True, type="primary"):
                st.session_state['vtm_user'] = user_choice
                save_log(user_choice, "시스템 접속")
                st.rerun()
    st.stop()

# ==================== 사이드바 ====================
with st.sidebar:
    st.markdown(f"### 👤 {st.session_state['vtm_user']}")
    
    # 본부장 전용 로그 확인 버튼
    if st.session_state['vtm_user'] == "박동진 본부장":
        st.markdown("---")
        st.markdown("### 🔐 관리자 전용")
        if st.button("📊 접속 기록 확인", use_container_width=True):
            st.session_state['show_logs'] = True
        st.markdown("---")
    
    st.markdown("---")
    
    selected_vtm = st.selectbox("📌 채널", list(CHANNELS.keys()))
    
    st.info(f"**📋 설명**\n\n{CHANNELS[selected_vtm]['desc']}")
    st.success(f"**🎯 목적**\n\n{CHANNELS[selected_vtm]['purpose']}")
    
    if st.button("🔥 떡상 분석 가동", use_container_width=True, type="primary"):
        st.session_state['trigger_analysis'] = True
        save_log(st.session_state['vtm_user'], f"{selected_vtm} 분석 실행")
    
    st.markdown("---")
    
    q_base, region = "", "KR"
    
    if selected_vtm == "브이티엠 엔터":
        theme = st.radio("🎵 테마", ["힙합(국내)", "AI 뮤비(전세계)"])
        config = CHANNELS[selected_vtm]['hiphop'] if "힙합" in theme else CHANNELS[selected_vtm]['ai']
        q_base, region = config['q'], config['region']
    elif selected_vtm == "시니어 스마일":
        theme = st.radio("🎤 테마", ["김복순(신나는)", "영원다방(정통)"])
        config = CHANNELS[selected_vtm]['boksun'] if "김복순" in theme else CHANNELS[selected_vtm]['youngwon']
        q_base, region = config['q'], config['region']
    else:
        q_base = CHANNELS[selected_vtm]['q']
        region = CHANNELS[selected_vtm]['region']
    
    st.markdown("---")
    
    content_type = st.radio("🎬 타입", ["미드폼만", "쇼츠만", "전체"])
    order_type = st.selectbox("📊 정렬", ["viewCount", "date"], 
                             format_func=lambda x: {"viewCount": "조회수순", "date": "최신순"}[x])
    period = st.selectbox("📅 기간", ["전체", "이번 주", "이번 달"])
    user_q = st.text_input("🔍 키워드")
    num_results = st.slider("📊 범위", 10, 30, 20)
    
    final_query = f"{q_base} {user_q}".strip()

# ==================== 본부장 전용: 로그 확인 ====================
if 'show_logs' in st.session_state and st.session_state['show_logs']:
    st.markdown("# 📊 시스템 접속 기록")
    st.markdown("---")
    
    all_logs = load_logs()
    
    if all_logs:
        # 통계
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("총 접속 횟수", len(all_logs))
        with col2:
            unique_users = len(set([log['user'] for log in all_logs]))
            st.metric("사용자 수", unique_users)
        with col3:
            today_logs = [log for log in all_logs if log['timestamp'].startswith(datetime.now().strftime('%Y-%m-%d'))]
            st.metric("오늘 접속", len(today_logs))
        
        st.markdown("---")
        
        # 로그 테이블
        df_logs = pd.DataFrame(all_logs)
        df_logs = df_logs.sort_values('timestamp', ascending=False)
        
        st.dataframe(df_logs, use_container_width=True, hide_index=True)
        
        st.markdown("---")
        
        # 엑셀 다운로드
        excel_logs = create_log_excel(all_logs)
        st.download_button(
            "📥 접속 기록 엑셀 다운로드",
            excel_logs,
            f"VTM_접속기록_{datetime.now().strftime('%Y%m%d')}.xlsx",
            use_container_width=True,
            type="primary"
        )
        
        if st.button("← 돌아가기", use_container_width=True):
            st.session_state['show_logs'] = False
            st.rerun()
    else:
        st.info("아직 접속 기록이 없습니다.")
        if st.button("← 돌아가기"):
            st.session_state['show_logs'] = False
            st.rerun()
    
    st.stop()

# ==================== 메인 ====================
st.markdown(f"<h1 style='text-align: center;'>🚀 {selected_vtm} 벤치마킹</h1>", unsafe_allow_html=True)
st.markdown("---")

# ==================== 분석 실행 ====================
if 'trigger_analysis' in st.session_state and st.session_state['trigger_analysis']:
    st.session_state['trigger_analysis'] = False
    
    with st.spinner('🎯 분석 중...'):
        data = fetch_youtube_native(final_query, region, order_type, period, content_type, num_results)
        
        if data and len(data) > 0:
            df = pd.DataFrame(data).sort_values("Viral Score", ascending=False).reset_index(drop=True)
            st.session_state['vtm_df'] = df
            st.success(f"✅ {len(df)}개 발굴")
            save_log(st.session_state['vtm_user'], f"{selected_vtm} 분석 완료: {len(df)}개 발굴")
        else:
            st.warning("⚠️ 데이터 없음")

# ==================== 결과 ====================
if st.session_state['vtm_df'] is not None and len(st.session_state['vtm_df']) > 0:
    df = st.session_state['vtm_df']
    
    tabs = st.tabs(["📊 그래프", "🏆 TOP 10 분석", "📄 상세 리포트", "🤖 AI 프롬프트"])
    
    # 탭 1: 그래프
    with tabs[0]:
        st.subheader("📊 Viral Score 분포")
        
        if PLOTLY_AVAILABLE:
            fig = px.scatter(df, x="구독자 수", y="Viral Score", size="조회수", color="채널명", hover_name="동영상 제목")
            fig.update_layout(height=600)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.dataframe(df[['채널명', '동영상 제목', 'Viral Score']])
    
    # 탭 2: TOP 10 분석
    with tabs[1]:
        st.markdown("## 🏆 TOP 10 경쟁 채널")
        
        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            rank_emoji = "🥇" if idx == 0 else "🥈" if idx == 1 else "🥉" if idx == 2 else f"**{idx+1}위**"
            
            with st.container():
                col1, col2 = st.columns([1, 4])
                
                with col1:
                    try:
                        st.image(row['썸네일'], use_container_width=True)
                    except:
                        st.write("🖼️")
                
                with col2:
                    st.markdown(f"### {rank_emoji} {row['동영상 제목']}")
                    st.markdown(f"**📺 채널**: {row['채널명']}")
                    st.markdown(f"**👥 구독자**: {row['구독자 수']:,}명 | **👁️ 조회수**: {row['조회수']:,}회")
                    st.markdown(f"**🔥 Viral Score**: {row['Viral Score']:,.0f}%")
                    
                    col_a, col_b = st.columns(2)
                    with col_a:
                        st.link_button("▶️ 영상 보기", row['영상 주소'], use_container_width=True)
                    with col_b:
                        st.link_button("📺 채널로 이동", row['채널 주소'], use_container_width=True)
                
                if idx < 3:
                    st.markdown(analyze_video(row, idx + 1, selected_vtm))
                
                st.markdown("---")
    
    # 탭 3: 상세 리포트
    with tabs[2]:
        st.subheader("📄 상세 리포트")
        
        display_df = df.copy()
        if len(display_df) > 0:
            display_df['동영상 제목'] = display_df['동영상 제목'].apply(lambda x: x[:50] + "..." if len(x) > 50 else x)
        
        st.dataframe(
            display_df,
            column_config={
                "썸네일": st.column_config.ImageColumn("썸네일", width="small"),
                "동영상 제목": st.column_config.TextColumn("제목", width="medium"),
                "영상 주소": st.column_config.LinkColumn("▶️"),
                "채널 주소": st.column_config.LinkColumn("🔗"),
                "Viral Score": st.column_config.NumberColumn("Viral Score", format="%.2f%%")
            },
            use_container_width=True,
            hide_index=True
        )
        
        st.markdown("---")
        
        excel_data = create_vtm_excel(df, "벤치마킹", st.session_state['vtm_user'])
        
        if st.download_button(
            "📥 엑셀 다운로드",
            excel_data,
            f"VTM_{selected_vtm}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            use_container_width=True,
            type="primary"
        ):
            save_log(st.session_state['vtm_user'], f"{selected_vtm} 엑셀 다운로드")
    
    # 탭 4: AI 프롬프트
    with tabs[3]:
        st.markdown("""
        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; border-radius: 15px;'>
            <h2 style='color: white; text-align: center;'>🤖 떡상 콘텐츠 제작 프롬프트</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        top3 = df.head(3) if len(df) >= 3 else df
        
        prompt = f"""# {selected_vtm} 떡상 콘텐츠 제작 프롬프트

## 📊 TOP 3 떡상 이유 요약

"""
        
        for idx, row in top3.iterrows():
            prompt += f"""
**{idx+1}위**: {row['동영상 제목']}
- 조회수: {row['조회수']:,}회 (Viral Score {row['Viral Score']:,.0f}%)
- 핵심: 구독자 {row['구독자 수']:,}명에서 {row['구독자 대비 조회']} 달성
"""
        
        prompt += f"""

## 🎯 미션
위 TOP 3 성공 요소를 분석하여 **{selected_vtm}**에 적용할 콘텐츠 기획안을 작성하세요.

## 📋 필수 산출물
1. 플레이리스트 제목 5개
2. 썸네일 디자인 가이드
3. 설명란 템플릿
4. 해시태그 30개
5. 업로드 전략

## 🎬 목표
Viral Score 1,000% 이상 달성
"""
        
        st.code(prompt, language="markdown")
        
        if st.button("📋 프롬프트 복사", use_container_width=True, type="primary"):
            st.success("✅ 텍스트를 마우스로 드래그하여 복사하세요!")

st.markdown("---")
st.markdown("<div style='text-align: center;'>⚖️ (주)브이티엠 (개발: 박동진 본부장)</div>", unsafe_allow_html=True)
